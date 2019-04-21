from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook(r'E:\to_hyf.xlsx')
print(wb.sheetnames)    # ['sheet1', 'sheet2', 'sheet3']


data = wb["sheet1"]
weight = wb["sheet2"]
result = wb["sheet3"]


print(data.title, data.max_row, data.max_column)    
print(weight.title, weight.max_row, weight.max_column)    
print(result.title, result.max_row, result.max_column)    

for c in range(1, 3):
    for r in range(1, data.max_row+1):
#         result_sheet1.write(r, c, data.cell(r,c).value)
        result.cell(row=r, column=c, value=data.cell(r,c).value)
    
    
for r in range(1, 3):
    for c in range(3, data.max_column+1):
        result.cell(row=r, column=c, value=data.cell(r,c).value)
        
        
for c in range(3, data.max_column+1):
    for r in range(3, data.max_row+1):
#         data.cell(r,c).value * weight.cell(0,c)
        result.cell(row=r, column=c, value=data.cell(r,c).value * weight.cell(1,c).value)
        
        
        
wb.save(r'E:\to_hyf.xlsx')
        
